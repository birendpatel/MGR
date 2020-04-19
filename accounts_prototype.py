"""
Name: Prototype for Accounts Reconciliation
Developer: Biren Patel
Date Created: 2020 January 02
Last Modified: 2020 January 02
Description: Accounts reconcilation on corporate bank statements
"""

import pandas as pd
from re import sub, match
from openpyxl import load_workbook, Workbook

def create_xlsx(data, filename='test.xlsx'):
    pd.DataFrame.to_excel(data, filename, index=False)

if __name__ == "__main__":
    #regular expression setup
    find_tx_tag = r'^TXEFILE'
    del_tx_tag_head = r'^TXEFILE\*0'
    del_tx_tag_tail = r'-0$'
    
    #read bank statement file
    bank_DF = pd.read_csv('chase_statement.csv')
    
    #strip txefile account ids from description into new column 
    bank_DF['Account ID'] = ""
    
    for i in range(bank_DF.shape[0]):
        if match(find_tx_tag, bank_DF.loc[i, ['Description']][0]):
            acct_id = bank_DF.loc[i, ['Description']][0]
            acct_id = sub(del_tx_tag_head, "", acct_id)
            acct_id = sub(del_tx_tag_tail, "", acct_id)
            
            if len(acct_id) != 8:
                raise ValueError("Account ID Less Than 8 Digits")
                
            bank_DF.loc[i, ['Account ID']] = acct_id
            
    #switch over to reconciliation file as read-only
    reco_rfile = load_workbook('reconciliation_report.xlsx', read_only=True)
    sheetname = reco_rfile.sheetnames[0]
    reco_rfile_ws = reco_rfile[sheetname]
    
    #a new workbook will store collected data from the read-only
    reco_wfile = Workbook()
    reco_wfile_ws = reco_wfile.active
    reco_wfile_ws.title = 'Reconciliation Data'
    
    #search first 50 rows to find the ID header in reconciliation file
    start_idx = 0
    
    for i in range(1,51):
        cell = 'A' + str(i)    
        if reco_rfile_ws[cell].value == 'ID':
            start_idx = i
            break
    else:
        raise IndexError("No ID Header Available in Reconciliation File")
        
    #prepare sheet in the new workbook
    reco_wfile_ws['A1'] = 'Account ID'
    reco_wfile_ws['B1'] = 'Cause Number'
    reco_wfile_ws['C1'] = 'Matter Number'
    reco_wfile_ws['D1'] = 'Reconciliation Description'
    reco_wfile_ws['E1'] = 'Submission Date'
    reco_wfile_ws['F1'] = 'Acceptance Date'
    
    #populate this sheet with data from read-only reconciliation file
    read_row = start_idx + 1
    write_row = 2
    
    while True:
        cell = 'A' + str(read_row)
        
        curr_val = reco_rfile_ws[cell].value
        
        if curr_val == None:
            break
        elif type(curr_val ) != str:
            TypeError("Fatal, reconciliation report next value at +3 not str")
        elif len(curr_val) != 8:
            ValueError("Fatal, reconciliation report account id not 8 digits")
        else:
            #add information at this position to the write sheet
            r_row = str(read_row)
            w_row = str(write_row)
            
            #account id
            reco_wfile_ws['A' + w_row] = curr_val
            #cause number
            reco_wfile_ws['B' + w_row] = reco_rfile_ws['D' + r_row].value
            #matter number
            reco_wfile_ws['C' + w_row] = reco_rfile_ws['E' + r_row].value
            #description
            jump = str(int(r_row) + 2)
            reco_wfile_ws['D' + w_row] = reco_rfile_ws['D' + jump].value
            #submission date
            reco_wfile_ws['E' + w_row] = reco_rfile_ws['B' + r_row].value
            #acceptance date
            reco_wfile_ws['F' + w_row] = reco_rfile_ws['C' + r_row].value
            
            
        read_row += 3
        write_row += 1
        
        #failsafe against infinite loop
        if read_row == 1000000:
            raise ValueError("Fatal, reconciliation report read row is large")
            
    #openpyxl workbook to pandas data frame
    reco_DF = pd.DataFrame(reco_wfile_ws.values)
    
    #replace header with first row
    reco_DF.columns = reco_DF.iloc[0]
    reco_DF = reco_DF.iloc[1:]
    
    #inner join bank data frame to reconciliation data frame
    data = bank_DF.merge(reco_DF, how='inner', on='Account ID')
