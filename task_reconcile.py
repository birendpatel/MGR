"""
Name: Accounts Reconciliation Task
Developer: Biren Patel
Date Created: 2020 January 02
Last Modified: 2020 January 13
Description: Accounts reconcilation on corporate bank statements. Linked to the
reconciliation task widget in the GUI application.
"""

import pandas as pd
from enum import Enum
from re import sub, match
from time import sleep
from openpyxl import load_workbook, Workbook
from PySide2.QtWidgets import QProgressBar

class status(Enum):
    NULL_STATUS = -1
    SUCCESS = 0
    FAIL_BANK_ACCT_LEN = 1
    FAIL_RECO_NO_ID = 2
    FAIL_RECO_NO_PLUS_3 = 3
    FAIL_RECO_ACCT_LEN = 4
    FAIL_RECO_LARGE = 5

def create_xlsx(data, filename='test.xlsx'):
    pd.DataFrame.to_excel(data, filename, index=False)

def reconcile(bank_path, reco_path, progress_bar):
    #create empty dataframe and status code to be returned
    data = pd.DataFrame({'NULL': []})
    status_code = status.NULL_STATUS

    #regular expression setup
    find_tx_tag = r'^TXEFILE'
    del_tx_tag_head = r'^TXEFILE\*0'
    del_tx_tag_tail = r'-0$'

    #read bank statement file and keep only required columns
    bank_DF = pd.read_csv(bank_path)
    bank_DF = bank_DF[['Transaction Date','Post Date','Description',\
                       'Amount','Memo']]

    #strip txefile account ids from description into new column
    progress_bar.setValue(10)
    bank_DF['Account ID'] = ""

    for i in range(bank_DF.shape[0]):
        if match(find_tx_tag, bank_DF.loc[i, ['Description']][0]):
            acct_id = bank_DF.loc[i, ['Description']][0]
            acct_id = sub(del_tx_tag_head, "", acct_id)
            acct_id = sub(del_tx_tag_tail, "", acct_id)

            if len(acct_id) != 8:
                status_code = status.FAIL_BANK_ACCT_LEN
                return (data, status_code)

            bank_DF.loc[i, ['Account ID']] = acct_id

    #switch over to reconciliation file as read-only
    progress_bar.setValue(15)
    reco_rfile = load_workbook(reco_path, read_only=True)
    sheetname = reco_rfile.sheetnames[0]
    reco_rfile_ws = reco_rfile[sheetname]

    #a new workbook will store collected data from the read-only
    reco_wfile = Workbook()
    reco_wfile_ws = reco_wfile.active
    reco_wfile_ws.title = 'Reconciliation Data'

    #search first 50 rows to find the ID header in reconciliation file
    progress_bar.setValue(20)
    start_idx = 0

    for i in range(1,51):
        cell = 'A' + str(i)
        if reco_rfile_ws[cell].value == 'ID':
            start_idx = i
            break
    else:
        #raise IndexError("No ID Header Available in Reconciliation File")
        status_code = status.FAIL_RECO_NO_ID
        return (data, status_code)

    #prepare sheet in the new workbook
    progress_bar.setValue(25)
    reco_wfile_ws['A1'] = 'Account ID'
    reco_wfile_ws['B1'] = 'Cause Number'
    reco_wfile_ws['C1'] = 'Matter Number'
    reco_wfile_ws['D1'] = 'Reconciliation Description'

    #populate this sheet with data from read-only reconciliation file
    progress_bar.setValue(30)
    read_row = start_idx + 1
    write_row = 2

    while True:
        cell = 'A' + str(read_row)

        curr_val = reco_rfile_ws[cell].value

        if curr_val == None:
            break
        elif type(curr_val) != str:
            status_code = status.FAIL_RECO_NO_PLUS_3
            return (data, status_code)
        elif len(curr_val) != 8:
            status_code = status.FAIL_RECO_ACCT_LEN
            return (data, status_code)
        else:
            #add information at this position to the write sheet
            r_row = str(read_row)
            w_row = str(write_row)

            #account id
            reco_wfile_ws['A' + w_row] = curr_val
            #cause number
            reco_wfile_ws['B' + w_row] = reco_rfile_ws['D' + r_row].value
            #matter number
            reco_wfile_ws['C' + w_row] = reco_rfile_ws['E' + r_row].value
            #description
            jump = str(int(r_row) + 2)
            reco_wfile_ws['D' + w_row] = reco_rfile_ws['D' + jump].value

        read_row += 3
        write_row += 1

        #update progress bar in calling widget
        val = progress_bar.value()
        if write_row % 5 == 0 and val < 90:
            progress_bar.setValue(val + 1)

        #failsafe against infinite loop
        if read_row == 1000000:
            status_code = status.FAIL_RECO_LARGE
            return (data, status_code)

    #openpyxl workbook to pandas data frame
    progress_bar.setValue(99)
    reco_DF = pd.DataFrame(reco_wfile_ws.values)

    #replace header with first row
    reco_DF.columns = reco_DF.iloc[0]
    reco_DF = reco_DF.iloc[1:]

    #inner join bank data frame to reconciliation data frame
    data = bank_DF.merge(reco_DF, how='left', on='Account ID')

    #all good, pass control back
    status_code = status.SUCCESS
    return (data, status_code)
